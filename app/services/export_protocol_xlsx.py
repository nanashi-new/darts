from __future__ import annotations

import os
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


@dataclass
class ProtocolData:
    tournament_name: str
    competition_title: str
    category: str
    format_type: str  # "classification" | "501" | "norms"
    date: str
    venue: str
    city: str
    org_name: str
    logo_path: str | None
    jury: list[dict[str, str]] = field(default_factory=list)
    # Each result dict may contain: place, fio, birth_year, coach,
    # score_set, score_sector20, score_big_round, points_total,
    # rank_achieved, region, current_rank (optional, player's rank before competition)
    results: list[dict[str, object]] = field(default_factory=list)


def export_protocol_xlsx(path: str, data: ProtocolData) -> None:
    """Export a formatted protocol to XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Протокол"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    current_row = 1

    # Logo
    if data.logo_path and os.path.isfile(data.logo_path):
        try:
            img = Image(data.logo_path)
            img.width = 100
            img.height = 80
            ws.add_image(img, "A1")
            current_row = 5
        except Exception:  # noqa: BLE001
            pass

    # Organization name (may contain newlines)
    org_lines = data.org_name.split("\n") if data.org_name else []
    for line in org_lines:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        cell = ws.cell(row=current_row, column=1, value=line.strip())
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

    # City - right aligned
    if data.city:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        cell = ws.cell(row=current_row, column=1, value=data.city)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        current_row += 1

    current_row += 1  # blank row

    # Competition title
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    cell = ws.cell(row=current_row, column=1, value=data.competition_title)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    # PROTOKOL REZULTATOV
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    cell = ws.cell(row=current_row, column=1, value="\u041f\u0420\u041e\u0422\u041e\u041a\u041e\u041b \u0420\u0415\u0417\u0423\u041b\u042c\u0422\u0410\u0422\u041e\u0412")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    # Category
    if data.category:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        cell = ws.cell(row=current_row, column=1, value=data.category)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

    # Format type label
    format_labels = {
        "classification": "\u041a\u043b\u0430\u0441\u0441\u0438\u0444\u0438\u043a\u0430\u0446\u0438\u044f",
        "501": "501 - \u043e\u0434\u0438\u043d\u043e\u0447\u043d\u044b\u0439 \u0440\u0430\u0437\u0440\u044f\u0434",
        "norms": "\u0421\u0434\u0430\u0447\u0430 \u043d\u043e\u0440\u043c\u0430\u0442\u0438\u0432\u043e\u0432",
    }
    format_label = format_labels.get(data.format_type, data.format_type)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    cell = ws.cell(row=current_row, column=1, value=format_label)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    # Venue + date
    venue_date = f"{data.venue}, {data.date}" if data.venue else data.date
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    cell = ws.cell(row=current_row, column=1, value=venue_date)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 2  # blank row after

    # Jury section
    if data.jury:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        cell = ws.cell(row=current_row, column=1, value="\u0421\u0443\u0434\u0435\u0439\u0441\u043a\u0430\u044f \u043a\u043e\u043b\u043b\u0435\u0433\u0438\u044f:")
        cell.font = Font(bold=True)
        current_row += 1

        for jury_member in data.jury:
            ws.cell(row=current_row, column=1, value=jury_member.get("position", ""))
            ws.cell(row=current_row, column=2, value=jury_member.get("name", ""))
            ws.cell(row=current_row, column=3, value=jury_member.get("category", ""))
            ws.cell(row=current_row, column=4, value=jury_member.get("city", ""))
            current_row += 1

        current_row += 1  # blank row after jury

    # Results table
    if data.format_type == "501":
        columns = [
            "\u041c\u0435\u0441\u0442\u043e",
            "\u0424\u0430\u043c\u0438\u043b\u0438\u044f, \u0418\u043c\u044f",
            "\u0413\u043e\u0434 \u0440\u043e\u0436\u0434\u0435\u043d\u0438\u044f",
            "\u0417\u0432\u0430\u043d\u0438\u0435, \u0440\u0430\u0437\u0440\u044f\u0434",
            "\u0421\u0443\u0431\u044a\u0435\u043a\u0442 \u0420\u0424, \u0433\u043e\u0440\u043e\u0434",
            "\u0422\u0440\u0435\u043d\u0435\u0440",
            "\u0412\u044b\u043f\u043e\u043b\u043d\u0435\u043d \u0440\u0430\u0437\u0440\u044f\u0434",
        ]
        result_keys = ["place", "fio", "birth_year", "current_rank", "region", "coach", "rank_achieved"]
    else:
        columns = [
            "\u041c\u0435\u0441\u0442\u043e",
            "\u0424\u0418\u041e",
            "\u0413/\u0420",
            "\u0422\u0440\u0435\u043d\u0435\u0440",
            "\u041d\u0430\u0431\u043e\u0440 \u043e\u0447\u043a\u043e\u0432",
            "\u0421\u0435\u043a\u0442\u043e\u0440 20",
            "\u0411\u043e\u043b\u044c\u0448\u043e\u0439 \u0440\u0430\u0443\u043d\u0434",
            "\u0418\u0442\u043e\u0433\u043e",
            "\u0412\u044b\u043f\u043e\u043b\u043d\u0435\u043d\u043d\u044b\u0439 \u0440\u0430\u0437\u0440\u044f\u0434",
        ]
        result_keys = [
            "place", "fio", "birth_year", "coach",
            "score_set", "score_sector20", "score_big_round",
            "points_total", "rank_achieved",
        ]

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    current_row += 1

    # Data rows
    for result in data.results:
        for col_idx, key in enumerate(result_keys, start=1):
            value = result.get(key, "")
            cell = ws.cell(row=current_row, column=col_idx, value=value if value is not None else "")
            cell.border = thin_border
            if col_idx == 1 or col_idx > 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

    current_row += 2  # blank rows before signatures

    # Signature section
    chief_judge_name = ""
    chief_secretary_name = ""
    for jury_member in data.jury:
        pos = (jury_member.get("position") or "").lower()
        if "\u0433\u043b\u0430\u0432\u043d\u044b\u0439 \u0441\u0443\u0434\u044c\u044f" in pos or "chief judge" in pos:
            chief_judge_name = jury_member.get("name", "")
        elif "\u0433\u043b\u0430\u0432\u043d\u044b\u0439 \u0441\u0435\u043a\u0440\u0435\u0442\u0430\u0440\u044c" in pos or "secretary" in pos:
            chief_secretary_name = jury_member.get("name", "")

    ws.cell(row=current_row, column=1, value=f"\u0413\u043b\u0430\u0432\u043d\u044b\u0439 \u0441\u0443\u0434\u044c\u044f _________ {chief_judge_name}")
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"\u0413\u043b\u0430\u0432\u043d\u044b\u0439 \u0441\u0435\u043a\u0440\u0435\u0442\u0430\u0440\u044c _________ {chief_secretary_name}")

    # Set column widths
    col_widths = [8, 25, 12, 20, 14, 14, 14, 12, 18]
    for idx, width in enumerate(col_widths, start=1):
        if idx <= len(columns):
            ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(path)
