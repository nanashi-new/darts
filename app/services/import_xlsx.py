from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
import json
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.db.database import get_default_database_path
from app.db.repositories import PlayerRepository, ResultRepository, TournamentRepository
from app.domain.points import points_for_place
from app.domain.ranks import calculate_points_classification
from app.services.norms_loader import load_norms_from_settings


@dataclass(frozen=True)
class ParsedTable:
    headers: list[str]
    rows: list[dict[str, object]]


@dataclass(frozen=True)
class ImportParseReport:
    headers: list[str]
    rows: list[dict[str, object]]
    errors: list[str]
    warnings: list[str]
    missing_required_columns: list[str]
    needs_mapping: bool
    confidence: float


@dataclass(frozen=True)
class TableBlock:
    sheet_name: str
    start_row: int
    end_row: int
    header_mapping: dict[str, str]
    rows: list[dict[str, object]]
    warnings: list[str]
    errors: list[str]
    needs_mapping: bool
    confidence: float
    missing_required_columns: list[str]


SUPPORTED_MAPPING_KEYS = (
    "fio",
    "birth_year",
    "birth_date",
    "place",
    "score_set",
    "score_sector20",
    "score_big_round",
)


@dataclass(frozen=True)
class ImportProfile:
    name: str
    required_columns: list[str]
    header_aliases: dict[str, list[str]]


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return "".join(ch for ch in text if ch.isalnum())


def detect_headers(row_values: Iterable[object]) -> dict[str, int]:
    synonyms = {
        "fio": ["фио", "игрок", "фамилияимя", "фамилия", "имя"],
        "birth": ["др", "датарождения", "годрождения", "рождения"],
        "coach": ["тренер", "coach"],
        "place": ["место", "place"],
        "score_set": ["набор", "очки", "наборочков", "score"],
        "score_sector20": ["с20", "sector20", "сектор20"],
        "score_big_round": ["бр", "biground", "большойраунд"],
    }
    normalized_synonyms = {
        key: {_normalize_header(item) for item in values}
        for key, values in synonyms.items()
    }

    mapping: dict[str, int] = {}
    for idx, cell_value in enumerate(row_values):
        normalized = _normalize_header(cell_value)
        if not normalized:
            continue
        for key, options in normalized_synonyms.items():
            if normalized in options:
                mapping[key] = idx
                break
    return mapping


def _default_required_fields() -> dict[str, str]:
    return {
        "fio": "ФИО",
        "place": "Место",
        "score_set": "Очки",
    }


def _calculate_mapping_stats(header_mapping: dict[str, int]) -> tuple[list[str], bool, float]:
    required_fields = _default_required_fields()
    missing_required = [label for key, label in required_fields.items() if key not in header_mapping]
    total_required = len(required_fields)
    found_required = total_required - len(missing_required)
    confidence = found_required / total_required if total_required else 0.0
    needs_mapping = confidence < 1.0
    return missing_required, needs_mapping, confidence


def _is_row_empty(row_values: Iterable[object]) -> bool:
    for value in row_values:
        if value is None:
            continue
        if str(value).strip() != "":
            return False
    return True


def _row_has_total(row_values: Iterable[object]) -> bool:
    for value in row_values:
        if value is None:
            continue
        if "итого" in str(value).strip().lower():
            return True
    return False


def _parse_first_table(path: str) -> tuple[
    list[str],
    list[dict[str, object]],
    dict[str, int],
    bool,
]:
    try:
        workbook = load_workbook(path, data_only=True)
    except (InvalidFileException, OSError):
        return [], [], {}, False
    sheet = workbook.active

    header_mapping: dict[str, int] = {}
    header_labels: list[str] = []
    rows: list[dict[str, object]] = []
    header_found = False

    for row in sheet.iter_rows(values_only=True):
        row_values = list(row)
        if not header_found:
            candidate_mapping = detect_headers(row_values)
            if candidate_mapping.get("fio") is not None:
                header_mapping = candidate_mapping
                header_labels = [str(value).strip() if value is not None else "" for value in row_values]
                header_found = True
            continue

        if _is_row_empty(row_values) or _row_has_total(row_values):
            break

        row_data: dict[str, object] = {
            "fio": None,
            "birth": None,
            "coach": None,
            "place": None,
            "score_set": None,
            "score_sector20": None,
            "score_big_round": None,
        }
        for key in row_data:
            if key in header_mapping:
                row_data[key] = row_values[header_mapping[key]]
        rows.append(row_data)

    return header_labels, rows, header_mapping, header_found


def _profile_storage_path() -> Path:
    return get_default_database_path().parent / "import_profiles.json"


def save_import_profile(profile: ImportProfile | dict[str, object]) -> None:
    payload = profile if isinstance(profile, dict) else {
        "name": profile.name,
        "required_columns": profile.required_columns,
        "header_aliases": profile.header_aliases,
    }
    name = str(payload.get("name", "")).strip()
    if not name:
        raise ValueError("У профиля должно быть имя.")

    required_columns = [str(item) for item in payload.get("required_columns", [])]
    aliases_raw = payload.get("header_aliases", {})
    if not isinstance(aliases_raw, dict):
        raise ValueError("header_aliases должен быть словарём.")

    header_aliases = {
        str(key): [str(alias) for alias in aliases]
        for key, aliases in aliases_raw.items()
        if isinstance(aliases, list)
    }

    profiles = list_import_profiles()
    filtered = [item for item in profiles if item.name != name]
    filtered.append(ImportProfile(name=name, required_columns=required_columns, header_aliases=header_aliases))

    path = _profile_storage_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps([
            {
                "name": p.name,
                "required_columns": p.required_columns,
                "header_aliases": p.header_aliases,
            }
            for p in filtered
        ], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def list_import_profiles() -> list[ImportProfile]:
    path = _profile_storage_path()
    if not path.exists():
        return []

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []

    profiles: list[ImportProfile] = []
    if not isinstance(raw, list):
        return profiles
    for item in raw:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()
        if not name:
            continue
        required_columns = [str(value) for value in item.get("required_columns", [])]
        aliases = item.get("header_aliases", {})
        if not isinstance(aliases, dict):
            aliases = {}
        header_aliases = {
            str(key): [str(alias) for alias in values]
            for key, values in aliases.items()
            if isinstance(values, list)
        }
        profiles.append(
            ImportProfile(
                name=name,
                required_columns=required_columns,
                header_aliases=header_aliases,
            )
        )
    return profiles




def delete_import_profile(name: str) -> None:
    target = name.strip()
    if not target:
        return
    profiles = [p for p in list_import_profiles() if p.name != target]
    path = _profile_storage_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps([
            {
                "name": p.name,
                "required_columns": p.required_columns,
                "header_aliases": p.header_aliases,
            }
            for p in profiles
        ], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
def apply_profile_to_headers(
    profile: ImportProfile | dict[str, object],
    headers_row: Iterable[object],
) -> tuple[dict[str, str], float]:
    aliases_raw = profile.header_aliases if isinstance(profile, ImportProfile) else profile.get("header_aliases", {})
    required_columns = profile.required_columns if isinstance(profile, ImportProfile) else profile.get("required_columns", [])
    if not isinstance(aliases_raw, dict):
        aliases_raw = {}

    normalized_headers = [str(value).strip() if value is not None else "" for value in headers_row]
    mapping: dict[str, str] = {}
    matched_required = 0
    required_set = {str(value) for value in required_columns}

    for header in normalized_headers:
        normalized = _normalize_header(header)
        if not normalized:
            continue
        for internal_key, aliases in aliases_raw.items():
            if not isinstance(aliases, list):
                continue
            normalized_aliases = {_normalize_header(alias) for alias in aliases}
            if normalized in normalized_aliases:
                mapping[header] = str(internal_key)
                if internal_key in required_set:
                    matched_required += 1
                break

    total_required = len(required_set)
    confidence = matched_required / total_required if total_required else 0.0
    return mapping, confidence


def _rows_for_table(
    sheet,
    start_index: int,
    header_mapping: dict[str, int],
    max_col: int,
) -> tuple[list[dict[str, object]], int]:
    rows: list[dict[str, object]] = []
    idx = start_index
    sheet_rows = list(sheet.iter_rows(values_only=True, max_col=max_col))
    while idx < len(sheet_rows):
        row_values = list(sheet_rows[idx])
        if _is_row_empty(row_values) or _row_has_total(row_values):
            break

        if detect_headers(row_values).get("fio") is not None:
            break

        row_data: dict[str, object] = {
            "fio": None,
            "birth": None,
            "coach": None,
            "place": None,
            "score_set": None,
            "score_sector20": None,
            "score_big_round": None,
        }
        for key in row_data:
            if key in header_mapping:
                row_data[key] = row_values[header_mapping[key]]
        rows.append(row_data)
        idx += 1

    return rows, idx


def read_table_block_preview(
    path: str,
    block: TableBlock,
    preview_rows: int = 8,
) -> tuple[list[str], list[list[object]]]:
    try:
        workbook = load_workbook(path, data_only=True)
    except (InvalidFileException, OSError):
        return [], []

    sheet = workbook[block.sheet_name] if block.sheet_name in workbook.sheetnames else workbook.active
    max_col = max(sheet.max_column, 1)
    sheet_rows = list(sheet.iter_rows(values_only=True, max_col=max_col))

    header_index = max(block.start_row - 1, 0)
    if header_index >= len(sheet_rows):
        return [], []

    headers = [str(value).strip() if value is not None else "" for value in list(sheet_rows[header_index])]
    data: list[list[object]] = []
    start_data = header_index + 1
    end_data = min(max(block.end_row - 1, start_data), len(sheet_rows) - 1)
    for idx in range(start_data, end_data + 1):
        data.append(list(sheet_rows[idx]))
        if len(data) >= max(preview_rows, 1):
            break
    return headers, data


def parse_table_block_with_mapping(
    path: str,
    block: TableBlock,
    column_mapping: dict[str, str],
) -> list[dict[str, object]]:
    headers, data_rows = read_table_block_preview(
        path,
        block,
        preview_rows=max(block.end_row - block.start_row, 1),
    )
    if not headers:
        return []

    header_to_index = {header: idx for idx, header in enumerate(headers)}
    rows: list[dict[str, object]] = []
    for values in data_rows:
        row_data: dict[str, object] = {
            "fio": None,
            "birth": None,
            "coach": None,
            "place": None,
            "score_set": None,
            "score_sector20": None,
            "score_big_round": None,
        }
        for internal_key, header in column_mapping.items():
            if internal_key not in SUPPORTED_MAPPING_KEYS:
                continue
            column_index = header_to_index.get(header)
            if column_index is None or column_index >= len(values):
                continue
            if internal_key in {"birth_year", "birth_date"}:
                row_data["birth"] = values[column_index]
            else:
                row_data[internal_key] = values[column_index]
        rows.append(row_data)
    return rows


def parse_tables_from_xlsx_with_report(path: str) -> list[TableBlock]:
    try:
        workbook = load_workbook(path, data_only=True)
    except (InvalidFileException, OSError):
        return []

    blocks: list[TableBlock] = []
    for sheet in workbook.worksheets:
        max_col = max(sheet.max_column, 1)
        sheet_rows = list(sheet.iter_rows(values_only=True, max_col=max_col))
        idx = 0
        while idx < len(sheet_rows):
            row_values = list(sheet_rows[idx])
            header_mapping = detect_headers(row_values)
            if header_mapping.get("fio") is None:
                idx += 1
                continue

            header_labels = [str(value).strip() if value is not None else "" for value in row_values]
            rows, end_idx = _rows_for_table(sheet, idx + 1, header_mapping, max_col)
            warnings = validate_rows(rows)
            missing_required, needs_mapping, confidence = _calculate_mapping_stats(header_mapping)

            if (confidence < 1.0 or needs_mapping) and header_labels:
                for profile in list_import_profiles():
                    profile_mapping, profile_confidence = apply_profile_to_headers(profile, header_labels)
                    if profile_confidence > confidence:
                        confidence = profile_confidence
                        needs_mapping = confidence < 1.0
                        missing_required = [
                            _default_required_fields().get(key, key)
                            for key in profile.required_columns
                            if key not in profile_mapping.values()
                        ]

            source_to_internal = {
                header_labels[column_idx]: key
                for key, column_idx in header_mapping.items()
                if 0 <= column_idx < len(header_labels)
            }

            errors = []
            for label in missing_required:
                errors.append(f"Не найден столбец {label}.")

            blocks.append(
                TableBlock(
                    sheet_name=sheet.title,
                    start_row=idx + 1,
                    end_row=end_idx,
                    header_mapping=source_to_internal,
                    rows=rows,
                    warnings=warnings,
                    errors=errors,
                    needs_mapping=needs_mapping,
                    confidence=confidence,
                    missing_required_columns=missing_required,
                )
            )
            idx = end_idx + 1

    return blocks


def import_batch_from_folder(folder: str, recursive: bool = False) -> dict[str, object]:
    base_path = Path(folder)
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    files = sorted(base_path.glob(pattern)) if base_path.exists() else []

    items: list[dict[str, object]] = []
    success = 0
    error = 0

    for file_path in files:
        try:
            tables = parse_tables_from_xlsx_with_report(str(file_path))
            if not tables:
                items.append(
                    {
                        "path": str(file_path),
                        "status": "error",
                        "message": "Не удалось распознать таблицы.",
                        "tables": 0,
                    }
                )
                error += 1
                continue

            item_status = "ok"
            message = "OK"
            if all((not block.rows) or block.errors for block in tables):
                item_status = "error"
                message = "; ".join(block.errors[0] for block in tables if block.errors) or "Нет данных."

            items.append(
                {
                    "path": str(file_path),
                    "status": item_status,
                    "message": message,
                    "tables": len(tables),
                }
            )
            if item_status == "ok":
                success += 1
            else:
                error += 1
        except Exception as exc:  # noqa: BLE001
            items.append(
                {
                    "path": str(file_path),
                    "status": "error",
                    "message": str(exc),
                    "tables": 0,
                }
            )
            error += 1

    return {
        "success": success,
        "error": error,
        "items": items,
    }


def parse_first_table_from_xlsx(path: str) -> tuple[list[str], list[dict[str, object]]]:
    header_labels, rows, _, _ = _parse_first_table(path)
    return header_labels, rows


def parse_first_table_from_xlsx_with_report(path: str) -> ImportParseReport:
    headers, rows, header_mapping, header_found = _parse_first_table(path)
    warnings = validate_rows(rows)
    errors: list[str] = []

    if not header_found:
        errors.append("Не найден заголовок таблицы.")

    missing_required, needs_mapping, confidence = _calculate_mapping_stats(header_mapping)
    for label in missing_required:
        errors.append(f"Не найден столбец {label}.")

    return ImportParseReport(
        headers=headers,
        rows=rows,
        errors=errors,
        warnings=warnings,
        missing_required_columns=missing_required,
        needs_mapping=needs_mapping,
        confidence=confidence,
    )


def _is_number(value: object) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.replace(",", "."))
        except ValueError:
            return False
        return True
    return False


def validate_rows(rows: Iterable[dict[str, object]]) -> list[str]:
    warnings: list[str] = []
    numeric_fields = {
        "place": "место",
        "score_set": "очки (набор)",
        "score_sector20": "сектор 20",
        "score_big_round": "большой раунд",
    }
    for idx, row in enumerate(rows, start=1):
        fio = row.get("fio")
        if fio is None or str(fio).strip() == "":
            warnings.append(f"Строка {idx}: пустое ФИО")
        for field, label in numeric_fields.items():
            value = row.get(field)
            if value is None or str(value).strip() == "":
                continue
            if parse_int(value, warnings) is None and not _is_number(value):
                warnings.append(
                    f"Строка {idx}: поле '{label}' не число ({value})"
                )
    return warnings


def _normalize_text(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_fio_key(value: object | None) -> str:
    text = _normalize_text(value).lower().replace("ё", "е")
    return " ".join(text.split())


def _birth_year_from_value(value: object | None) -> str | None:
    if value is None:
        return None
    text = _normalize_text(value)
    if len(text) >= 4 and text[:4].isdigit():
        return text[:4]
    return None


def _parse_fio(value: object) -> tuple[str, str, str | None]:
    text = _normalize_fio_key(value)
    parts = [part for part in text.split() if part]
    if not parts:
        return "", "", None
    last_name = parts[0]
    first_name = parts[1] if len(parts) > 1 else ""
    middle_name = " ".join(parts[2:]) if len(parts) > 2 else None
    return last_name, first_name, middle_name


def _parse_birth_value(value: object | None) -> tuple[str | None, str | None]:
    if value is None or _normalize_text(value) == "":
        return None, None
    if isinstance(value, datetime):
        return value.date().isoformat(), str(value.year)
    if isinstance(value, date):
        return value.isoformat(), str(value.year)
    if isinstance(value, (int, float)):
        year = int(value)
        if 1900 <= year <= 2100:
            return str(year), str(year)
        return None, None
    text = _normalize_text(value)
    if text.isdigit() and len(text) == 4:
        return text, text
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(text, fmt).date()
        except ValueError:
            continue
        return parsed.isoformat(), str(parsed.year)
    return None, None


def _player_match_rules_path() -> Path:
    return get_default_database_path().parent / "player_match_rules.json"


def _load_player_match_rules() -> dict[str, int]:
    path = _player_match_rules_path()
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(raw, dict):
        return {}
    rules: dict[str, int] = {}
    for key, player_id in raw.items():
        try:
            rules[str(key)] = int(player_id)
        except (TypeError, ValueError):
            continue
    return rules


def _save_player_match_rules(rules: dict[str, int]) -> None:
    path = _player_match_rules_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(rules, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _player_match_key(fio: object | None, birth_date_or_year: object | None) -> str:
    birth_date, birth_year = _parse_birth_value(birth_date_or_year)
    birth_token = birth_date or birth_year or ""
    return f"{_normalize_fio_key(fio)}|{birth_token}"


def find_player_candidates(
    fio: object,
    birth_date_or_year: object | None,
    *,
    player_repo: PlayerRepository,
) -> list[dict[str, object]]:
    fio_key = _normalize_fio_key(fio)
    if not fio_key:
        return []

    input_birth_date, input_birth_year = _parse_birth_value(birth_date_or_year)
    candidates: list[dict[str, object]] = []
    for player in player_repo.list():
        player_fio = _normalize_fio_key(
            " ".join(
                part for part in (
                    player.get("last_name"),
                    player.get("first_name"),
                    player.get("middle_name"),
                ) if part
            )
        )
        if player_fio != fio_key:
            continue

        if input_birth_date or input_birth_year:
            player_birth_raw = player.get("birth_date")
            player_birth_text = _normalize_text(player_birth_raw)
            player_birth_year = _birth_year_from_value(player_birth_raw)
            has_birth_match = False

            if input_birth_date and player_birth_text == input_birth_date:
                has_birth_match = True
            if input_birth_year and player_birth_year == input_birth_year:
                has_birth_match = True

            if not has_birth_match:
                continue

        candidates.append(player)
    return candidates


def _parse_integer_value(value: object | None) -> tuple[int | None, bool]:
    if value is None or _normalize_text(value) == "":
        return None, False
    if isinstance(value, bool):
        return None, False
    if isinstance(value, int):
        return value, False
    if isinstance(value, float):
        if value.is_integer():
            return int(value), False
        return None, True
    text = _normalize_text(value).replace(",", ".")
    try:
        decimal_value = Decimal(text)
    except InvalidOperation:
        return None, False
    if decimal_value != decimal_value.to_integral_value():
        return None, True
    return int(decimal_value), False


def parse_int(value: object | None, warnings: list[str] | None = None) -> int | None:
    parsed, has_fraction = _parse_integer_value(value)
    if has_fraction:
        if warnings is not None:
            warnings.append(f"некорректное целое число: {value}")
        return None
    return parsed


def import_tournament_rows(
    *,
    connection,
    rows: Iterable[dict[str, object]],
    tournament_name: str,
    tournament_date: str | None,
    category_code: str | None,
    source_files: list[str] | None = None,
    player_match_resolver: Callable[[str, str | None, list[dict[str, object]]], dict[str, object] | None] | None = None,
) -> tuple[int, bool]:
    tournament_repo = TournamentRepository(connection)
    player_repo = PlayerRepository(connection)
    result_repo = ResultRepository(connection)
    norms_load = load_norms_from_settings()
    norms, norms_loaded = norms_load.norms, norms_load.loaded

    tournament_id = tournament_repo.create(
        {
            "name": tournament_name,
            "date": tournament_date,
            "category_code": category_code,
            "league_code": None,
            "source_files": json.dumps(source_files or []),
        }
    )
    remembered_rules = _load_player_match_rules()

    for row in rows:
        fio = row.get("fio")
        if fio is None or _normalize_text(fio) == "":
            continue
        last_name, first_name, middle_name = _parse_fio(fio)
        birth_date, birth_year = _parse_birth_value(row.get("birth"))

        candidates = find_player_candidates(
            fio=fio,
            birth_date_or_year=birth_date or birth_year,
            player_repo=player_repo,
        )

        player: dict[str, object] | None = None
        if len(candidates) == 1:
            player = candidates[0]
        elif len(candidates) > 1:
            match_key = _player_match_key(fio, birth_date or birth_year)
            remembered_player_id = remembered_rules.get(match_key)
            if remembered_player_id is not None:
                player = next((item for item in candidates if int(item["id"]) == remembered_player_id), None)

            if player is None:
                if player_match_resolver is None:
                    raise ValueError(f"Найдено несколько игроков для '{fio}'.")
                resolution = player_match_resolver(str(fio), birth_date or birth_year, candidates)
                if not resolution:
                    raise ValueError("Импорт отменён пользователем.")
                action = str(resolution.get("action") or "cancel")
                if action == "cancel":
                    raise ValueError("Импорт отменён пользователем.")
                if action == "select":
                    selected_player_id = int(resolution.get("player_id"))
                    player = next((item for item in candidates if int(item["id"]) == selected_player_id), None)
                    if player is None:
                        raise ValueError("Выбранный игрок отсутствует в списке кандидатов.")
                    if bool(resolution.get("remember")):
                        remembered_rules[match_key] = selected_player_id
                        _save_player_match_rules(remembered_rules)
                elif action != "create":
                    raise ValueError("Неизвестное решение по выбору игрока.")

        if player is None:
            player_id = player_repo.create(
                {
                    "last_name": last_name,
                    "first_name": first_name,
                    "middle_name": middle_name,
                    "birth_date": birth_date,
                    "gender": None,
                    "coach": _normalize_text(row.get("coach")) or None,
                    "club": None,
                    "notes": None,
                }
            )
        else:
            player_id = int(player["id"])

        place = parse_int(row.get("place"))
        score_set = parse_int(row.get("score_set"))
        score_sector20 = parse_int(row.get("score_sector20"))
        score_big_round = parse_int(row.get("score_big_round"))

        gender = None if player is None else player.get("gender")
        ranks, points_classification = calculate_points_classification(
            score_set=score_set,
            score_sector20=score_sector20,
            score_big_round=score_big_round,
            gender=gender,
            birth_date=birth_date,
            tournament_date=tournament_date,
            norms=norms or {},
        )
        points_place = points_for_place(place) if place is not None else 0
        points_total = points_place + points_classification

        result_repo.create(
            {
                "tournament_id": tournament_id,
                "player_id": player_id,
                "place": place,
                "score_set": score_set,
                "score_sector20": score_sector20,
                "score_big_round": score_big_round,
                "rank_set": ranks["rank_set"],
                "rank_sector20": ranks["rank_sector20"],
                "rank_big_round": ranks["rank_big_round"],
                "points_classification": points_classification,
                "points_place": points_place,
                "points_total": points_total,
                "calc_version": "v2",
            }
        )

    return tournament_id, norms_loaded


def import_tournament_results(
    *,
    connection,
    file_path: str,
    tournament_name: str,
    tournament_date: str | None,
    category_code: str | None,
    player_match_resolver: Callable[[str, str | None, list[dict[str, object]]], dict[str, object] | None] | None = None,
) -> tuple[int, bool]:
    _, rows = parse_first_table_from_xlsx(file_path)
    if not rows:
        raise ValueError("Не удалось найти таблицу в файле.")

    return import_tournament_rows(
        connection=connection,
        rows=rows,
        tournament_name=tournament_name,
        tournament_date=tournament_date,
        category_code=category_code,
        source_files=[file_path],
        player_match_resolver=player_match_resolver,
    )
