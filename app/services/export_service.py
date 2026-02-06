from __future__ import annotations

import os
from datetime import date
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from PySide6.QtWidgets import QTableView


class ExportService:
    @staticmethod
    def _escape_pdf_text(text: str) -> str:
        return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    def _write_fallback_pdf(
        self,
        path: str,
        header_lines: list[str],
        columns: Sequence[str],
        rows: Sequence[Sequence[str]],
    ) -> None:
        lines: list[str] = []
        lines.extend(str(item) for item in header_lines)
        if columns:
            lines.append(" | ".join(str(item) for item in columns))
        for row in rows:
            lines.append(" | ".join("" if value is None else str(value) for value in row))

        stream_lines = ["BT", "/F1 11 Tf", "50 800 Td", "14 TL"]
        for idx, line in enumerate(lines[:120]):
            prefix = "" if idx == 0 else "T* "
            stream_lines.append(f"{prefix}({self._escape_pdf_text(line)}) Tj")
        stream_lines.append("ET")
        stream = "\n".join(stream_lines).encode("latin-1", errors="replace")

        objects = [
            b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
            b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n",
            b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 842 595] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n",
            b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n",
            f"5 0 obj << /Length {len(stream)} >> stream\n".encode("ascii") + stream + b"\nendstream endobj\n",
        ]

        pdf = bytearray(b"%PDF-1.4\n")
        offsets = [0]
        for obj in objects:
            offsets.append(len(pdf))
            pdf.extend(obj)
        xref_offset = len(pdf)
        pdf.extend(f"xref\n0 {len(offsets)}\n".encode("ascii"))
        pdf.extend(b"0000000000 65535 f \n")
        for offset in offsets[1:]:
            pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
        pdf.extend(
            f"trailer << /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
        )
        Path(path).write_bytes(pdf)

    def export_table_pdf(
        self,
        table: "QTableView",
        path: str,
        header_lines: Iterable[str],
    ) -> None:
        columns, rows = self._extract_table_data(table)
        column_widths = [max(table.columnWidth(i), 80) for i in range(len(columns))]
        self.export_dataset_pdf(path, header_lines, columns, rows, column_widths)

    def print_table(
        self,
        table: "QTableView",
        parent,
        header_lines: Iterable[str],
    ) -> bool:
        from PySide6.QtGui import QPageSize
        from PySide6.QtPrintSupport import QPrintDialog, QPrinter

        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageOrientation(QPrinter.Landscape)
        printer.setPageSize(QPageSize(QPageSize.A4))
        dialog = QPrintDialog(printer, parent)
        if dialog.exec() != QPrintDialog.Accepted:
            return False

        columns, rows = self._extract_table_data(table)
        column_widths = [max(table.columnWidth(i), 80) for i in range(len(columns))]
        self._render_dataset_to_printer(printer, list(header_lines), columns, rows, column_widths)
        return True

    def export_table_xlsx(
        self,
        table: "QTableView",
        path: str,
        header_lines: Iterable[str],
    ) -> None:
        columns, rows = self._extract_table_data(table)
        self.export_dataset_xlsx(path, header_lines, columns, rows)

    def save_table_image(self, table: "QTableView", path: str, full_table: bool = False) -> None:
        extension = Path(path).suffix.lower()
        image_format = "JPG" if extension in {".jpg", ".jpeg"} else "PNG"
        if not full_table:
            pixmap = table.viewport().grab()
            if not pixmap.save(path, image_format):
                raise OSError("Не удалось сохранить изображение.")
            return

        columns, rows = self._extract_table_data(table)
        column_widths = [max(table.columnWidth(i), 80) for i in range(len(columns))]
        self.export_dataset_image(path, columns, rows, column_widths)

    def export_dataset(
        self,
        export_format: str,
        path: str,
        header_lines: Iterable[str],
        columns: Sequence[str],
        rows: Sequence[Sequence[str]],
        column_widths: Sequence[int] | None = None,
        full_image: bool = True,
    ) -> None:
        normalized = export_format.lower()
        if normalized == "pdf":
            self.export_dataset_pdf(path, header_lines, columns, rows, column_widths)
        elif normalized == "xlsx":
            self.export_dataset_xlsx(path, header_lines, columns, rows)
        elif normalized in {"png", "jpg", "jpeg"}:
            if not full_image:
                raise ValueError("Режим видимой области доступен только для экспорта из таблицы UI.")
            self.export_dataset_image(path, columns, rows, column_widths)
        else:
            raise ValueError(f"Неподдерживаемый формат: {export_format}")

    def export_dataset_pdf(
        self,
        path: str,
        header_lines: Iterable[str],
        columns: Sequence[str],
        rows: Sequence[Sequence[str]],
        column_widths: Sequence[int] | None = None,
    ) -> None:
        header_lines_list = list(header_lines)
        rows_list = [["" if value is None else str(value) for value in row] for row in rows]
        try:
            from PySide6.QtGui import QPageSize
            from PySide6.QtPrintSupport import QPrinter

            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(path)
            printer.setPageOrientation(QPrinter.Landscape)
            printer.setPageSize(QPageSize(QPageSize.A4))
            self._render_dataset_to_printer(
                printer,
                header_lines_list,
                list(columns),
                rows_list,
                list(column_widths) if column_widths is not None else [120] * len(columns),
            )
            if not os.path.exists(path):
                raise OSError("Не удалось сохранить PDF файл.")
            return
        except Exception:  # noqa: BLE001
            self._write_fallback_pdf(path, header_lines_list, columns, rows_list)
            if not os.path.exists(path):
                raise OSError("Не удалось сохранить PDF файл.")

    def export_dataset_xlsx(
        self,
        path: str,
        header_lines: Iterable[str],
        columns: Sequence[str],
        rows: Sequence[Sequence[str]],
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Экспорт"

        header_lines_list = list(header_lines)
        current_row = 1
        for line in header_lines_list:
            sheet.cell(row=current_row, column=1, value=line)
            current_row += 1

        header_row = current_row
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for column, header_text in enumerate(columns, start=1):
            cell = sheet.cell(row=header_row, column=column, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill

        current_row += 1
        alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for row in rows:
            for column, value in enumerate(row, start=1):
                cell = sheet.cell(row=current_row, column=column, value=value)
                cell.alignment = alignment
            current_row += 1

        sheet.freeze_panes = f"A{header_row + 1}"

        for column_index in range(1, len(columns) + 1):
            max_length = len(str(columns[column_index - 1]))
            for row_index in range(header_row + 1, current_row):
                value = sheet.cell(row=row_index, column=column_index).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)

        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        workbook.save(path)

    def export_dataset_image(
        self,
        path: str,
        columns: Sequence[str],
        rows: Sequence[Sequence[str]],
        column_widths: Sequence[int] | None = None,
    ) -> None:
        if not columns:
            raise OSError("Нет данных для экспорта изображения.")
        widths = list(column_widths) if column_widths else [140] * len(columns)
        padding = 6
        row_height = 28
        header_height = 34
        width = sum(widths) + 1
        height = header_height + row_height * len(rows) + 1
        from PySide6.QtCore import QRect, Qt
        from PySide6.QtGui import QImage, QPainter

        image = QImage(width, height, QImage.Format_ARGB32)
        image.fill(Qt.white)

        painter = QPainter(image)
        try:
            painter.setPen(Qt.black)
            x = 0
            for index, title in enumerate(columns):
                rect = QRect(x, 0, widths[index], header_height)
                painter.fillRect(rect, Qt.lightGray)
                painter.drawRect(rect)
                painter.drawText(rect.adjusted(padding, 0, -padding, 0), Qt.AlignVCenter | Qt.AlignLeft, title)
                x += widths[index]

            for row_index, row in enumerate(rows):
                y = header_height + row_index * row_height
                x = 0
                for col_index, value in enumerate(row):
                    rect = QRect(x, y, widths[col_index], row_height)
                    painter.drawRect(rect)
                    painter.drawText(
                        rect.adjusted(padding, 0, -padding, 0),
                        Qt.AlignVCenter | Qt.AlignLeft,
                        "" if value is None else str(value),
                    )
                    x += widths[col_index]
        finally:
            painter.end()

        extension = Path(path).suffix.lower()
        image_format = "JPG" if extension in {".jpg", ".jpeg"} else "PNG"
        if not image.save(path, image_format):
            raise OSError("Не удалось сохранить изображение.")

    def _extract_table_data(self, table: "QTableView") -> tuple[list[str], list[list[str]]]:
        model = table.model()
        if model is None:
            return [], []
        from PySide6.QtCore import Qt

        columns = [str(model.headerData(column, Qt.Horizontal) or "") for column in range(model.columnCount())]
        rows: list[list[str]] = []
        for row in range(model.rowCount()):
            rows.append([
                "" if model.index(row, column).data() is None else str(model.index(row, column).data())
                for column in range(model.columnCount())
            ])
        return columns, rows

    def _render_dataset_to_printer(
        self,
        printer,
        header_lines: list[str],
        columns: list[str],
        rows: list[list[str]],
        column_widths: list[int],
    ) -> None:
        if not columns:
            return
        from PySide6.QtCore import QRect, Qt
        from PySide6.QtGui import QPainter, QTextOption
        from PySide6.QtPrintSupport import QPrinter

        painter = QPainter(printer)
        try:
            page_rect = printer.pageRect(QPrinter.DevicePixel).adjusted(40, 40, -40, -40)
            total_width = sum(column_widths)
            scaled_widths = [max(60, int(page_rect.width() * width / total_width)) for width in column_widths]
            scaled_widths[-1] += page_rect.width() - sum(scaled_widths)

            header_font = painter.font()
            header_font.setPointSize(12)
            table_header_font = painter.font()
            table_header_font.setPointSize(10)
            table_header_font.setBold(True)
            body_font = painter.font()
            body_font.setPointSize(9)
            text_option = QTextOption()
            text_option.setWrapMode(QTextOption.WordWrap)

            def draw_page_header() -> tuple[int, int]:
                y = page_rect.top()
                painter.setFont(header_font)
                metrics = painter.fontMetrics()
                for line in header_lines:
                    line_rect = QRect(page_rect.left(), y, page_rect.width(), metrics.height())
                    painter.drawText(line_rect, Qt.AlignLeft | Qt.AlignVCenter, line)
                    y += metrics.height() + 4

                header_row_height = 30
                painter.setFont(table_header_font)
                x = page_rect.left()
                for index, title in enumerate(columns):
                    rect = QRect(x, y, scaled_widths[index], header_row_height)
                    painter.fillRect(rect, Qt.lightGray)
                    painter.drawRect(rect)
                    painter.drawText(rect.adjusted(6, 0, -6, 0), Qt.AlignVCenter | Qt.AlignLeft, title)
                    x += scaled_widths[index]
                return y + header_row_height, page_rect.bottom()

            y, bottom = draw_page_header()
            painter.setFont(body_font)
            body_metrics = painter.fontMetrics()

            for row in rows:
                row_height = 24
                for index, value in enumerate(row):
                    cell_rect = body_metrics.boundingRect(
                        QRect(0, 0, scaled_widths[index] - 12, 10_000),
                        Qt.TextWordWrap,
                        "" if value is None else str(value),
                    )
                    row_height = max(row_height, cell_rect.height() + 10)

                if y + row_height > bottom:
                    printer.newPage()
                    y, bottom = draw_page_header()
                    painter.setFont(body_font)

                x = page_rect.left()
                for index, value in enumerate(row):
                    rect = QRect(x, y, scaled_widths[index], row_height)
                    painter.drawRect(rect)
                    painter.drawText(rect.adjusted(6, 4, -6, -4), "" if value is None else str(value), text_option)
                    x += scaled_widths[index]
                y += row_height
        finally:
            painter.end()

    @staticmethod
    def format_date_label() -> str:
        return date.today().strftime("%d.%m.%Y")
