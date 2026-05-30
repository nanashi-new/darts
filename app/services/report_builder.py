"""Configurable report builder service."""

from __future__ import annotations

import json
import os
import sqlite3
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.analytics import AnalyticsService
from app.services.export_service import ExportService


@dataclass(frozen=True)
class ReportConfig:
    sections: list[str] = field(default_factory=list)
    period_start: str | None = None
    period_end: str | None = None
    league_code: str | None = None
    category_code: str | None = None
    player_ids: list[int] | None = None
    output_format: str = "text"

    def to_json(self) -> str:
        return json.dumps(
            {
                "sections": self.sections,
                "period_start": self.period_start,
                "period_end": self.period_end,
                "league_code": self.league_code,
                "category_code": self.category_code,
                "player_ids": self.player_ids,
                "output_format": self.output_format,
            },
            ensure_ascii=False,
        )

    @staticmethod
    def from_json(data: str) -> "ReportConfig":
        obj = json.loads(data)
        return ReportConfig(
            sections=obj.get("sections", []),
            period_start=obj.get("period_start"),
            period_end=obj.get("period_end"),
            league_code=obj.get("league_code"),
            category_code=obj.get("category_code"),
            player_ids=obj.get("player_ids"),
            output_format=obj.get("output_format", "text"),
        )


@dataclass(frozen=True)
class ReportResult:
    file_path: str
    sections_generated: list[str]
    total_rows: int


class ReportBuilderService:
    """Builds configurable reports from DB data."""

    def __init__(self) -> None:
        self._analytics = AnalyticsService()
        self._export = ExportService()

    def build_report(
        self,
        connection: sqlite3.Connection,
        config: ReportConfig,
        output_dir: str,
    ) -> ReportResult:
        sections_data: dict[str, tuple[list[str], list[list[str]]]] = {}
        total_rows = 0

        for section in config.sections:
            if section == "rating":
                columns, rows = self._build_rating_section(connection, config)
            elif section == "tournaments":
                columns, rows = self._build_tournaments_section(connection, config)
            elif section == "players":
                columns, rows = self._build_players_section(connection, config)
            elif section == "analytics":
                columns, rows = self._build_analytics_section(connection, config)
            else:
                continue
            sections_data[section] = (columns, rows)
            total_rows += len(rows)

        os.makedirs(output_dir, exist_ok=True)
        fmt = config.output_format.lower()
        if fmt == "xlsx":
            file_path = os.path.join(output_dir, "report.xlsx")
            self._write_xlsx(file_path, sections_data)
        elif fmt == "pdf":
            file_path = os.path.join(output_dir, "report.pdf")
            self._write_pdf(file_path, sections_data)
        else:
            file_path = os.path.join(output_dir, "report.txt")
            self._write_text(file_path, sections_data)

        return ReportResult(
            file_path=file_path,
            sections_generated=list(sections_data.keys()),
            total_rows=total_rows,
        )

    def _build_rating_section(
        self, connection: sqlite3.Connection, config: ReportConfig
    ) -> tuple[list[str], list[list[str]]]:
        clauses: list[str] = []
        params: list[object] = []
        if config.league_code:
            clauses.append("t.league_code = ?")
            params.append(config.league_code)
        if config.category_code:
            clauses.append("t.category_code = ?")
            params.append(config.category_code)
        if config.period_start:
            clauses.append("t.date >= ?")
            params.append(config.period_start)
        if config.period_end:
            clauses.append("t.date <= ?")
            params.append(config.period_end)
        where_sql = ""
        if clauses:
            where_sql = "WHERE " + " AND ".join(clauses)
        rows = connection.execute(
            f"""
            SELECT
                p.last_name || ' ' || p.first_name AS fio,
                r.place,
                r.points_total,
                t.name AS tournament_name,
                t.date
            FROM results r
            JOIN players p ON p.id = r.player_id
            JOIN tournaments t ON t.id = r.tournament_id
            {where_sql}
            ORDER BY r.points_total DESC, r.place ASC
            """,
            params,
        ).fetchall()
        columns = ["Игрок", "Место", "Очки", "Турнир", "Дата"]
        data = [[str(c) if c is not None else "" for c in row] for row in rows]
        return columns, data

    def _build_tournaments_section(
        self, connection: sqlite3.Connection, config: ReportConfig
    ) -> tuple[list[str], list[list[str]]]:
        clauses: list[str] = []
        params: list[object] = []
        if config.period_start:
            clauses.append("date >= ?")
            params.append(config.period_start)
        if config.period_end:
            clauses.append("date <= ?")
            params.append(config.period_end)
        if config.league_code:
            clauses.append("league_code = ?")
            params.append(config.league_code)
        if config.category_code:
            clauses.append("category_code = ?")
            params.append(config.category_code)
        where_sql = ""
        if clauses:
            where_sql = "WHERE " + " AND ".join(clauses)
        rows = connection.execute(
            f"""
            SELECT name, date, category_code, league_code, status
            FROM tournaments
            {where_sql}
            ORDER BY date DESC, name
            """,
            params,
        ).fetchall()
        columns = ["Название", "Дата", "Категория", "Лига", "Статус"]
        data = [[str(c) if c is not None else "" for c in row] for row in rows]
        return columns, data

    def _build_players_section(
        self, connection: sqlite3.Connection, config: ReportConfig
    ) -> tuple[list[str], list[list[str]]]:
        if config.player_ids:
            placeholders = ", ".join("?" for _ in config.player_ids)
            rows = connection.execute(
                f"""
                SELECT last_name, first_name, middle_name, birth_date, club
                FROM players
                WHERE id IN ({placeholders})
                ORDER BY last_name, first_name
                """,
                config.player_ids,
            ).fetchall()
        else:
            rows = connection.execute(
                "SELECT last_name, first_name, middle_name, birth_date, club FROM players ORDER BY last_name, first_name"
            ).fetchall()
        columns = ["Фамилия", "Имя", "Отчество", "Дата рождения", "Клуб"]
        data = [[str(c) if c is not None else "" for c in row] for row in rows]
        return columns, data

    def _build_analytics_section(
        self, connection: sqlite3.Connection, config: ReportConfig
    ) -> tuple[list[str], list[list[str]]]:
        top = self._analytics.top_results(
            connection,
            period_start=config.period_start,
            period_end=config.period_end,
            limit=10,
        )
        columns = ["Игрок", "Турнир", "Дата", "Очки", "Место"]
        data = [
            [entry.fio, entry.tournament_name, entry.date, str(entry.points_total), str(entry.place)]
            for entry in top
        ]
        return columns, data

    def _write_text(
        self,
        path: str,
        sections_data: dict[str, tuple[list[str], list[list[str]]]],
    ) -> None:
        lines: list[str] = []
        for section_name, (columns, rows) in sections_data.items():
            lines.append(f"=== {section_name.upper()} ===")
            lines.append(" | ".join(columns))
            lines.append("-" * 60)
            for row in rows:
                lines.append(" | ".join(row))
            lines.append("")
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    def _write_xlsx(
        self,
        path: str,
        sections_data: dict[str, tuple[list[str], list[list[str]]]],
    ) -> None:
        workbook = Workbook()
        first = True
        for section_name, (columns, rows) in sections_data.items():
            if first:
                sheet = workbook.active
                assert sheet is not None
                sheet.title = section_name
                first = False
            else:
                sheet = workbook.create_sheet(title=section_name)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            for col_idx, col_name in enumerate(columns, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = header_fill
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        workbook.save(path)

    def _write_pdf(
        self,
        path: str,
        sections_data: dict[str, tuple[list[str], list[list[str]]]],
    ) -> None:
        all_columns: list[str] = []
        all_rows: list[list[str]] = []
        header_lines: list[str] = ["Отчет Darts Liga"]
        for section_name, (columns, rows) in sections_data.items():
            if not all_columns:
                all_columns = columns
                all_rows = list(rows)
            else:
                all_rows.append([""] * len(all_columns))
                all_rows.append([f"--- {section_name} ---"] + [""] * (len(all_columns) - 1))
                for row in rows:
                    padded = row + [""] * (len(all_columns) - len(row))
                    all_rows.append(padded[:len(all_columns)])
        self._export._write_fallback_pdf(path, header_lines, all_columns, all_rows)
