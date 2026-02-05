from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Iterable

from openpyxl import load_workbook

from app.domain.points import points_for_rank


@dataclass(frozen=True)
class RankThreshold:
    score_min: int
    rank: str


Norms = dict[str, dict[str, dict[str, list[RankThreshold]]]]


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return "".join(ch for ch in text if ch.isalnum())


def _detect_header_mapping(row_values: Iterable[object]) -> dict[str, int]:
    synonyms = {
        "gender": ["пол", "gender", "sex"],
        "age_group": ["возраст", "возрастнаягруппа", "agegroup", "age"],
        "discipline": ["дисциплина", "discipline", "вид"],
        "score_min": ["минимум", "минимальныйрезультат", "scoremin", "порог"],
        "rank": ["разряд", "rank", "квалификация"],
    }
    normalized_synonyms = {
        key: {_normalize_header(item) for item in values}
        for key, values in synonyms.items()
    }
    mapping: dict[str, int] = {}
    for idx, cell_value in enumerate(row_values):
        normalized = _normalize_header(cell_value)
        if not normalized:
            continue
        for key, options in normalized_synonyms.items():
            if normalized in options:
                mapping[key] = idx
                break
    return mapping


def _normalize_gender(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"m", "м", "male", "муж", "мужской"}:
        return "M"
    if text in {"f", "ж", "female", "жен", "женский"}:
        return "F"
    return None


def _normalize_age_group(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    normalized = "".join(ch for ch in text if ch.isalnum())
    if "u10" in normalized or "до10" in normalized:
        return "U10"
    if "u12" in normalized or "до12" in normalized:
        return "U12"
    if "u15" in normalized or "до15" in normalized:
        return "U15"
    if "u18" in normalized or "до18" in normalized:
        return "U18"
    digits = [int("".join(group)) for group in _extract_digit_groups(normalized)]
    if digits:
        max_age = max(digits)
        if max_age <= 10:
            return "U10"
        if max_age <= 12:
            return "U12"
        if max_age <= 15:
            return "U15"
        return "U18"
    return None


def _extract_digit_groups(text: str) -> list[list[str]]:
    groups: list[list[str]] = []
    current: list[str] = []
    for ch in text:
        if ch.isdigit():
            current.append(ch)
        elif current:
            groups.append(current)
            current = []
    if current:
        groups.append(current)
    return groups


def _normalize_discipline(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    normalized = "".join(ch for ch in text if ch.isalnum())
    if normalized in {"set", "набор", "наборочков", "очки", "score"}:
        return "SET"
    if normalized in {"sector20", "сектор20", "с20", "c20"}:
        return "SECTOR20"
    if normalized in {"biground", "большойраунд", "бр"}:
        return "BIGROUND"
    return None


def _normalize_rank(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower().replace(" ", "")
    if text in {"3юн", "3юнош", "3юношеский"}:
        return "3юн"
    if text in {"2юн", "2юнош", "2юношеский"}:
        return "2юн"
    if text in {"1юн", "1юнош", "1юношеский"}:
        return "1юн"
    if text in {"3сп", "3спорт", "3спортивный"}:
        return "3сп"
    if text in {"2сп", "2спорт", "2спортивный"}:
        return "2сп"
    if text in {"1сп", "1спорт", "1спортивный"}:
        return "1сп"
    if text in {"кмс"}:
        return "КМС"
    return None


def load_norms_from_xlsx(path: str) -> Norms:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    norms: Norms = {}

    header_mapping: dict[str, int] = {}
    header_found = False

    for row in sheet.iter_rows(values_only=True):
        row_values = list(row)
        if not header_found:
            candidate = _detect_header_mapping(row_values)
            if {"gender", "age_group", "discipline", "score_min", "rank"}.issubset(
                candidate
            ):
                header_mapping = candidate
                header_found = True
            continue

        if not any(value is not None and str(value).strip() for value in row_values):
            break

        gender = _normalize_gender(row_values[header_mapping["gender"]])
        age_group = _normalize_age_group(row_values[header_mapping["age_group"]])
        discipline = _normalize_discipline(row_values[header_mapping["discipline"]])
        rank = _normalize_rank(row_values[header_mapping["rank"]])
        score_raw = row_values[header_mapping["score_min"]]

        if gender is None or age_group is None or discipline is None or rank is None:
            continue
        if score_raw is None:
            continue
        try:
            score_min = int(float(str(score_raw).replace(",", ".")))
        except ValueError:
            continue

        norms.setdefault(gender, {}).setdefault(age_group, {}).setdefault(
            discipline, []
        ).append(RankThreshold(score_min=score_min, rank=rank))

    for gender_data in norms.values():
        for age_data in gender_data.values():
            for discipline, thresholds in age_data.items():
                age_data[discipline] = sorted(
                    thresholds, key=lambda item: item.score_min
                )

    return norms


def _parse_date(value: object | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit() and len(text) == 4:
        return date(int(text), 1, 1)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def get_age_group(birth_date: object | None, tournament_date: object | None) -> str | None:
    birth = _parse_date(birth_date)
    tournament = _parse_date(tournament_date)
    if birth is None or tournament is None:
        return None
    age = tournament.year - birth.year - (
        (tournament.month, tournament.day) < (birth.month, birth.day)
    )
    if age < 0:
        return None
    if age < 10:
        return "U10"
    if age < 12:
        return "U12"
    if age < 15:
        return "U15"
    return "U18"


def get_rank(
    score: object | None,
    gender: object | None,
    birth_date: object | None,
    tournament_date: object | None,
    discipline: object | None,
    norms: Norms,
) -> str | None:
    if score is None:
        return None
    if not norms:
        return None
    try:
        score_value = int(float(str(score).replace(",", ".")))
    except ValueError:
        return None

    gender_key = _normalize_gender(gender)
    discipline_key = _normalize_discipline(discipline)
    age_group = get_age_group(birth_date, tournament_date)
    if gender_key is None or discipline_key is None or age_group is None:
        return None

    thresholds = norms.get(gender_key, {}).get(age_group, {}).get(discipline_key)
    if not thresholds:
        return None

    best_rank: str | None = None
    for threshold in thresholds:
        if isinstance(threshold, RankThreshold):
            score_min = threshold.score_min
            rank = threshold.rank
        else:
            score_min, rank = threshold
        if score_value >= score_min:
            best_rank = rank
        else:
            break
    return best_rank


def rank_points(rank: str | None) -> int:
    return points_for_rank(rank)


def calculate_points_classification(
    *,
    score_set: object | None,
    score_sector20: object | None,
    score_big_round: object | None,
    gender: object | None,
    birth_date: object | None,
    tournament_date: object | None,
    norms: Norms,
) -> tuple[dict[str, str | None], int]:
    ranks = {"rank_set": None, "rank_sector20": None, "rank_big_round": None}
    if not norms:
        return ranks, 0

    age_group = get_age_group(birth_date, tournament_date)
    if age_group is None:
        return ranks, 0

    disciplines = [
        ("SET", "rank_set", score_set),
        ("SECTOR20", "rank_sector20", score_sector20),
        ("BIGROUND", "rank_big_round", score_big_round),
    ]
    if age_group == "U10":
        disciplines = [("SET", "rank_set", score_set)]

    total_points = 0
    for discipline, key, score in disciplines:
        rank = get_rank(
            score,
            gender,
            birth_date,
            tournament_date,
            discipline,
            norms,
        )
        ranks[key] = rank
        total_points += rank_points(rank)

    return ranks, total_points
