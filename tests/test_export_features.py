from pathlib import Path

from openpyxl import load_workbook

from app.db.database import get_connection
from app.db.repositories import PlayerRepository, ResultRepository, TournamentRepository
from app.services.batch_export import BatchExportService
from app.services.export_service import ExportService


def _seed_sample_data(connection) -> tuple[int, int]:
    players = PlayerRepository(connection)
    tournaments = TournamentRepository(connection)
    results = ResultRepository(connection)

    player_id = players.create(
        {
            "last_name": "Иванов",
            "first_name": "Иван",
            "middle_name": "Иванович",
            "birth_date": "2000-01-01",
            "gender": "M",
            "coach": None,
            "club": None,
            "notes": None,
        }
    )
    tournament_id = tournaments.create(
        {
            "name": "Кубок города",
            "date": "2025-01-10",
            "category_code": "U12-M",
            "league_code": "A",
            "source_files": "[]",
        }
    )
    results.create(
        {
            "tournament_id": tournament_id,
            "player_id": player_id,
            "place": 1,
            "score_set": 100,
            "score_sector20": 30,
            "score_big_round": 70,
            "rank_set": "I",
            "rank_sector20": "I",
            "rank_big_round": "I",
            "points_classification": 20,
            "points_place": 40,
            "points_total": 60,
            "calc_version": "v1",
        }
    )
    return player_id, tournament_id


def test_export_dataset_xlsx_contains_headers(tmp_path: Path) -> None:
    path = tmp_path / "rating.xlsx"
    service = ExportService()
    service.export_dataset_xlsx(
        str(path),
        header_lines=["Рейтинг", "Дата: 01.01.2025"],
        columns=["Место", "ФИО", "Очки"],
        rows=[["1", "Очень Длинное ФИО", "123"]],
    )

    assert path.exists()
    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet.cell(row=3, column=1).value == "Место"
    assert sheet.cell(row=3, column=2).value == "ФИО"
    assert sheet.cell(row=3, column=3).value == "Очки"


def test_batch_export_creates_folder_and_files(tmp_path: Path) -> None:
    db_path = tmp_path / "app.db"
    connection = get_connection(db_path)
    _seed_sample_data(connection)

    service = BatchExportService(connection)
    result = service.export_all(tmp_path, export_format="xlsx", n_value=6)

    assert result.run_directory.exists()
    assert len(result.files_created) == 2
    assert all(path.exists() for path in result.files_created)
