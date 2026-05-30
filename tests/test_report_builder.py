"""Tests for report builder service and report templates."""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import pytest

from app.db.repositories import ReportTemplateRepository
from app.db.schema import initialize_schema
from app.services.report_builder import ReportBuilderService, ReportConfig, ReportResult


def _create_db(tmp_path: Path) -> sqlite3.Connection:
    db_path = str(tmp_path / "test.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    initialize_schema(conn)
    return conn


def _seed_data(conn: sqlite3.Connection) -> None:
    conn.execute(
        "INSERT INTO players (id, last_name, first_name) VALUES (1, 'Иванов', 'Иван')"
    )
    conn.execute(
        "INSERT INTO players (id, last_name, first_name) VALUES (2, 'Петров', 'Петр')"
    )
    conn.execute(
        "INSERT INTO tournaments (id, name, date, category_code, league_code, status) "
        "VALUES (1, 'Турнир А', '2024-03-01', 'U12', 'L1', 'published')"
    )
    conn.execute(
        "INSERT INTO tournaments (id, name, date, category_code, league_code, status) "
        "VALUES (2, 'Турнир Б', '2024-06-15', 'U14', 'L2', 'published')"
    )
    conn.execute(
        "INSERT INTO results (tournament_id, player_id, place, points_total) VALUES (1, 1, 1, 100)"
    )
    conn.execute(
        "INSERT INTO results (tournament_id, player_id, place, points_total) VALUES (1, 2, 2, 80)"
    )
    conn.execute(
        "INSERT INTO results (tournament_id, player_id, place, points_total) VALUES (2, 1, 2, 90)"
    )
    conn.execute(
        "INSERT INTO results (tournament_id, player_id, place, points_total) VALUES (2, 2, 1, 95)"
    )
    conn.commit()


class TestReportBuilderService:
    def test_build_text_report_with_rating(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        _seed_data(conn)
        service = ReportBuilderService()
        config = ReportConfig(sections=["rating"], output_format="text")
        output_dir = str(tmp_path / "output")
        result = service.build_report(conn, config, output_dir)

        assert result.file_path.endswith("report.txt")
        assert "rating" in result.sections_generated
        assert result.total_rows == 4
        content = Path(result.file_path).read_text(encoding="utf-8")
        assert "Иванов" in content
        assert "Петров" in content

    def test_report_with_period_filter(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        _seed_data(conn)
        service = ReportBuilderService()
        config = ReportConfig(
            sections=["tournaments"],
            period_start="2024-04-01",
            period_end="2024-12-31",
            output_format="text",
        )
        output_dir = str(tmp_path / "output")
        result = service.build_report(conn, config, output_dir)

        assert result.total_rows == 1
        content = Path(result.file_path).read_text(encoding="utf-8")
        assert "Турнир Б" in content
        assert "Турнир А" not in content

    def test_report_with_category_filter(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        _seed_data(conn)
        service = ReportBuilderService()
        config = ReportConfig(
            sections=["rating"],
            category_code="U12",
            output_format="text",
        )
        output_dir = str(tmp_path / "output")
        result = service.build_report(conn, config, output_dir)

        assert result.total_rows == 2
        content = Path(result.file_path).read_text(encoding="utf-8")
        assert "Турнир А" in content
        assert "Турнир Б" not in content

    def test_xlsx_report_generates_valid_file(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        _seed_data(conn)
        service = ReportBuilderService()
        config = ReportConfig(
            sections=["rating", "tournaments"],
            output_format="xlsx",
        )
        output_dir = str(tmp_path / "output")
        result = service.build_report(conn, config, output_dir)

        assert result.file_path.endswith("report.xlsx")
        assert Path(result.file_path).exists()
        assert Path(result.file_path).stat().st_size > 0

        from openpyxl import load_workbook

        wb = load_workbook(result.file_path)
        assert "rating" in wb.sheetnames
        assert "tournaments" in wb.sheetnames

    def test_pdf_report_generates_file(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        _seed_data(conn)
        service = ReportBuilderService()
        config = ReportConfig(
            sections=["rating"],
            output_format="pdf",
        )
        output_dir = str(tmp_path / "output")
        result = service.build_report(conn, config, output_dir)

        assert result.file_path.endswith("report.pdf")
        assert Path(result.file_path).exists()
        content = Path(result.file_path).read_bytes()
        assert content.startswith(b"%PDF")

    def test_analytics_section(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        _seed_data(conn)
        service = ReportBuilderService()
        config = ReportConfig(sections=["analytics"], output_format="text")
        output_dir = str(tmp_path / "output")
        result = service.build_report(conn, config, output_dir)

        assert "analytics" in result.sections_generated
        assert result.total_rows > 0

    def test_players_section_with_ids(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        _seed_data(conn)
        service = ReportBuilderService()
        config = ReportConfig(
            sections=["players"],
            player_ids=[1],
            output_format="text",
        )
        output_dir = str(tmp_path / "output")
        result = service.build_report(conn, config, output_dir)

        assert result.total_rows == 1
        content = Path(result.file_path).read_text(encoding="utf-8")
        assert "Иванов" in content
        assert "Петров" not in content


class TestReportTemplateRepository:
    def test_save_and_list(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        repo = ReportTemplateRepository(conn)
        tid = repo.save_template("Test Template", '{"sections":["rating"]}')
        assert tid > 0

        templates = repo.list_templates()
        assert len(templates) == 1
        assert templates[0]["name"] == "Test Template"
        assert templates[0]["config_json"] == '{"sections":["rating"]}'

    def test_get_template(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        repo = ReportTemplateRepository(conn)
        tid = repo.save_template("My Report", '{"sections":["tournaments"]}')

        row = repo.get_template(tid)
        assert row is not None
        assert row["name"] == "My Report"
        assert row["id"] == tid

    def test_delete_template(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        repo = ReportTemplateRepository(conn)
        tid = repo.save_template("To Delete", '{}')
        repo.delete_template(tid)

        assert repo.get_template(tid) is None
        assert len(repo.list_templates()) == 0

    def test_list_templates_ordered_by_created_at(self, tmp_path: Path) -> None:
        conn = _create_db(tmp_path)
        repo = ReportTemplateRepository(conn)
        repo.save_template("First", '{}')
        # Insert second with a later timestamp
        conn.execute(
            "INSERT INTO report_templates (name, config_json, created_at) VALUES (?, ?, ?)",
            ("Second", "{}", "2099-01-01 00:00:00"),
        )
        conn.commit()

        templates = repo.list_templates()
        assert len(templates) == 2
        assert templates[0]["name"] == "Second"
        assert templates[1]["name"] == "First"


class TestReportConfigSerialization:
    def test_to_json_and_from_json(self) -> None:
        config = ReportConfig(
            sections=["rating", "tournaments"],
            period_start="2024-01-01",
            period_end="2024-12-31",
            league_code="L1",
            category_code="U12",
            player_ids=[1, 2, 3],
            output_format="xlsx",
        )
        json_str = config.to_json()
        restored = ReportConfig.from_json(json_str)
        assert restored.sections == config.sections
        assert restored.period_start == config.period_start
        assert restored.period_end == config.period_end
        assert restored.league_code == config.league_code
        assert restored.category_code == config.category_code
        assert restored.player_ids == config.player_ids
        assert restored.output_format == config.output_format

    def test_from_json_defaults(self) -> None:
        json_str = json.dumps({"sections": ["players"]})
        config = ReportConfig.from_json(json_str)
        assert config.sections == ["players"]
        assert config.period_start is None
        assert config.period_end is None
        assert config.output_format == "text"
