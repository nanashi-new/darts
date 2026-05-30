import os
import struct
import zlib

import pytest

from app.settings import get_organization_profile, update_organization_profile


pytestmark = pytest.mark.integration


def test_organization_profile_crud(tmp_path, monkeypatch):
    """Test organization profile get/update cycle."""
    settings_file = tmp_path / "settings.json"
    monkeypatch.setattr("app.settings.get_settings_path", lambda: settings_file)

    profile = get_organization_profile()
    assert profile["org_name"] == ""
    assert profile["jury_members"] == []

    update_organization_profile({
        "org_name": "Test Org\nLine2",
        "city": "Moscow",
        "logo_path": None,
        "jury_members": [{"position": "Judge", "name": "Ivanov", "category": "VK", "city": "Moscow"}],
        "default_venue": "Arena",
    })

    profile = get_organization_profile()
    assert profile["org_name"] == "Test Org\nLine2"
    assert profile["city"] == "Moscow"
    assert len(profile["jury_members"]) == 1


def test_export_protocol_xlsx_creates_file(tmp_path):
    """Test XLSX protocol export creates a valid file."""
    from app.services.export_protocol_xlsx import ProtocolData, export_protocol_xlsx

    data = ProtocolData(
        tournament_name="Test Cup",
        competition_title="Regional Competition",
        category="Men",
        format_type="classification",
        date="2025-01-15",
        venue="Arena",
        city="Tver",
        org_name="Federation\nLine2",
        logo_path=None,
        jury=[{"position": "Chief Judge", "name": "Ivanov A.", "category": "VK", "city": "Tver"}],
        results=[
            {"place": 1, "fio": "Petrov Ivan", "birth_year": "2000", "coach": "Sidorov",
             "score_set": 100, "score_sector20": 30, "score_big_round": 70,
             "points_total": 200, "rank_achieved": "I", "region": "Tver"},
        ],
    )

    path = str(tmp_path / "protocol.xlsx")
    export_protocol_xlsx(path, data)

    assert os.path.exists(path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    # Verify some content exists
    found_protocol = False
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        for cell in row:
            if cell and "\u041f\u0420\u041e\u0422\u041e\u041a\u041e\u041b" in str(cell):
                found_protocol = True
                break
    assert found_protocol, "Header 'PROTOKOL' not found in XLSX"


def _create_minimal_png(path):
    """Create a minimal 1x1 white PNG file."""
    sig = b'\x89PNG\r\n\x1a\n'
    ihdr_data = struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0)
    ihdr_crc = zlib.crc32(b'IHDR' + ihdr_data) & 0xffffffff
    ihdr = struct.pack('>I', 13) + b'IHDR' + ihdr_data + struct.pack('>I', ihdr_crc)
    raw = b'\x00\xff\xff\xff'
    compressed = zlib.compress(raw)
    idat_crc = zlib.crc32(b'IDAT' + compressed) & 0xffffffff
    idat = struct.pack('>I', len(compressed)) + b'IDAT' + compressed + struct.pack('>I', idat_crc)
    iend_crc = zlib.crc32(b'IEND') & 0xffffffff
    iend = struct.pack('>I', 0) + b'IEND' + struct.pack('>I', iend_crc)
    with open(path, 'wb') as f:
        f.write(sig + ihdr + idat + iend)


def test_export_protocol_xlsx_with_logo(tmp_path):
    """Test XLSX protocol export with a logo image."""
    from app.services.export_protocol_xlsx import ProtocolData, export_protocol_xlsx

    logo_path = str(tmp_path / "logo.png")
    _create_minimal_png(logo_path)

    data = ProtocolData(
        tournament_name="Test",
        competition_title="Competition",
        category="Women",
        format_type="501",
        date="2025-02-20",
        venue="Hall",
        city="Moscow",
        org_name="Org",
        logo_path=logo_path,
        jury=[],
        results=[{"place": 1, "fio": "Ivanova M.", "birth_year": "1995", "coach": "",
                  "score_set": 0, "score_sector20": 0, "score_big_round": 0,
                  "points_total": 501, "rank_achieved": "", "region": "Moscow"}],
    )

    path = str(tmp_path / "protocol_logo.xlsx")
    export_protocol_xlsx(path, data)
    assert os.path.exists(path)


def test_export_protocol_docx_creates_file(tmp_path):
    """Test DOCX protocol export creates a valid file."""
    from app.services.export_protocol_docx import export_protocol_docx
    from app.services.export_protocol_xlsx import ProtocolData

    data = ProtocolData(
        tournament_name="Docx Test",
        competition_title="Cup of Region",
        category="Juniors",
        format_type="classification",
        date="2025-03-10",
        venue="Sports Center",
        city="SPb",
        org_name="Organization Name",
        logo_path=None,
        jury=[{"position": "Secretary", "name": "Petrov B.", "category": "I", "city": "SPb"}],
        results=[
            {"place": 1, "fio": "Smirnov A.", "birth_year": "2005", "coach": "Coach A",
             "score_set": 80, "score_sector20": 25, "score_big_round": 55,
             "points_total": 160, "rank_achieved": "II", "region": "SPb"},
            {"place": 2, "fio": "Kozlov B.", "birth_year": "2006", "coach": "Coach B",
             "score_set": 70, "score_sector20": 20, "score_big_round": 50,
             "points_total": 140, "rank_achieved": "III", "region": "SPb"},
        ],
    )

    path = str(tmp_path / "protocol.docx")
    export_protocol_docx(path, data)

    assert os.path.exists(path)
    from docx import Document
    doc = Document(path)
    text = "\n".join(p.text for p in doc.paragraphs)
    assert "\u041f\u0420\u041e\u0422\u041e\u041a\u041e\u041b" in text


def test_export_protocol_dialog_smoke():
    """Smoke test: ExportProtocolDialog can be instantiated."""
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    try:
        from PySide6.QtWidgets import QApplication
        if QApplication.instance() is None:
            QApplication([])
    except Exception:
        pytest.skip("PySide6 unavailable")

    from app.ui.export_protocol_dialog import ExportProtocolDialog

    tournament = {"name": "Test Tournament", "category_code": "U15-M", "date": "2025-01-01", "id": 1}
    results = [{"place": 1, "last_name": "Ivanov", "first_name": "Ivan", "middle_name": "",
                "birth_date": "2000-01-01", "score_set": 100, "score_sector20": 30,
                "score_big_round": 70, "points_total": 200}]

    dialog = ExportProtocolDialog(tournament=tournament, results=results)
    assert dialog.windowTitle() == "\u042d\u043a\u0441\u043f\u043e\u0440\u0442 \u043f\u0440\u043e\u0442\u043e\u043a\u043e\u043b\u0430"
    # Verify key widgets exist
    assert dialog._protocol_type_combo is not None
    assert dialog._export_btn is not None
