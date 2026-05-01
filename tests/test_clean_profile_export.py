from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.db.database import get_connection
from app.db.repositories import PlayerRepository, ResultRepository, TournamentRepository
from app.runtime_paths import get_runtime_paths
from app.services.batch_export import BatchExportService


pytestmark = pytest.mark.release_smoke


def _seed_export_data(connection) -> None:
    players = PlayerRepository(connection)
    tournaments = TournamentRepository(connection)
    results = ResultRepository(connection)

    player_id = players.create(
        {
            "last_name": "Иванов",
            "first_name": "Иван",
            "middle_name": None,
            "birth_date": "2012-01-02",
            "gender": "M",
            "coach": None,
            "club": None,
            "notes": None,
        }
    )
    tournament_id = tournaments.create(
        {
            "name": "Clean Profile Cup",
            "date": "2026-04-26",
            "category_code": "U12-M",
            "source_files": "[]",
            "status": "published",
            "has_draft_changes": 0,
        }
    )
    results.create(
        {
            "tournament_id": tournament_id,
            "player_id": player_id,
            "place": 1,
            "score_set": 120,
            "score_sector20": 45,
            "score_big_round": 78,
            "rank_set": None,
            "rank_sector20": None,
            "rank_big_round": None,
            "points_classification": 0,
            "points_place": 14,
            "points_total": 14,
            "calc_version": "v3_no_classification",
        }
    )


def test_clean_profile_exports_pdf_and_xlsx_to_profile_exports(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setenv("DARTS_PROFILE_ROOT", str(tmp_path / "profile"))
    paths = get_runtime_paths()
    connection = get_connection(paths.db_path)
    _seed_export_data(connection)

    service = BatchExportService(connection)
    pdf_result = service.export_all_to_profile(export_format="pdf", n_value=3)
    xlsx_result = service.export_all_to_profile(export_format="xlsx", n_value=3)

    assert pdf_result.run_directory.parent == paths.exports_dir
    assert xlsx_result.run_directory.parent == paths.exports_dir
    assert "exports" not in pdf_result.run_directory.relative_to(paths.exports_dir).parts
    assert "exports" not in xlsx_result.run_directory.relative_to(paths.exports_dir).parts

    pdf_files = [path for path in pdf_result.files_created if path.suffix.lower() == ".pdf"]
    xlsx_files = [path for path in xlsx_result.files_created if path.suffix.lower() == ".xlsx"]
    assert len(pdf_files) == 2
    assert len(xlsx_files) == 2

    for path in [*pdf_files, *xlsx_files]:
        assert path.exists()
        assert path.stat().st_size > 0
        assert path.is_relative_to(paths.exports_dir)

    assert pdf_files[0].read_bytes().startswith(b"%PDF")
    workbook = load_workbook(xlsx_files[0])
    assert workbook.active.cell(row=1, column=1).value in {"Рейтинг", "Протокол турнира"}
