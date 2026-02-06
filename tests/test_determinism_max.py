from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.db.database import get_connection
from app.db.repositories import PlayerRepository, ResultRepository, TournamentRepository
from app.services.batch_export import BatchExportService
from app.services.recalculate_tournament import recalculate_all_tournaments


def _seed_dataset(connection) -> None:
    players = PlayerRepository(connection)
    tournaments = TournamentRepository(connection)
    results = ResultRepository(connection)

    player_ids = []
    for i in range(10):
        player_ids.append(
            players.create(
                {
                    "last_name": f"Игроков{i}",
                    "first_name": "Тест",
                    "middle_name": None,
                    "birth_date": "2010-01-01",
                    "gender": "M",
                    "coach": None,
                    "club": None,
                    "notes": None,
                }
            )
        )

    tournament_ids = []
    for j in range(5):
        tournament_ids.append(
            tournaments.create(
                {
                    "name": f"Tour {j}",
                    "date": f"2025-01-{j + 1:02d}",
                    "category_code": "U12",
                    "league_code": None,
                    "source_files": "[]",
                }
            )
        )

    for t_idx, tournament_id in enumerate(tournament_ids):
        for p_idx, player_id in enumerate(player_ids):
            points_total = (10 - p_idx) * 5 + t_idx
            results.create(
                {
                    "tournament_id": tournament_id,
                    "player_id": player_id,
                    "place": p_idx + 1,
                    "score_set": 100 - p_idx,
                    "score_sector20": 20 + p_idx,
                    "score_big_round": 30 + p_idx,
                    "rank_set": None,
                    "rank_sector20": None,
                    "rank_big_round": None,
                    "points_classification": 1,
                    "points_place": points_total,
                    "points_total": points_total,
                    "calc_version": "v1",
                }
            )


def _xlsx_values(path: Path) -> list[tuple[object, ...]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    out: list[tuple[object, ...]] = []
    for row in ws.iter_rows(values_only=True):
        out.append(tuple(row))
    return out


def test_determinism_max(tmp_path: Path) -> None:
    connection = get_connection(tmp_path / "determinism.db")
    _seed_dataset(connection)

    report_first = recalculate_all_tournaments(connection=connection)
    export_first = BatchExportService(connection).export_all(tmp_path / "r1", export_format="xlsx", n_value=6)

    report_second = recalculate_all_tournaments(connection=connection)
    export_second = BatchExportService(connection).export_all(tmp_path / "r2", export_format="xlsx", n_value=6)

    assert report_first.results_updated == report_second.results_updated

    rows_first = connection.execute(
        "SELECT tournament_id, player_id, points_total FROM results ORDER BY tournament_id, points_total DESC, player_id"
    ).fetchall()
    rows_second = connection.execute(
        "SELECT tournament_id, player_id, points_total FROM results ORDER BY tournament_id, points_total DESC, player_id"
    ).fetchall()
    assert [tuple(r) for r in rows_first] == [tuple(r) for r in rows_second]

    first_rating = sorted([p for p in export_first.files_created if "rating_" in p.name])[0]
    second_rating = sorted([p for p in export_second.files_created if "rating_" in p.name])[0]

    assert _xlsx_values(first_rating) == _xlsx_values(second_rating)
